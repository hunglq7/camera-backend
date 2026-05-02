from datetime import datetime, timezone
from io import BytesIO
from typing import Tuple

import openpyxl
from models.camera import Camera
from utils.network import check_camera_status


def create_camera(db, camera_data):
    db_camera = Camera(
        name=camera_data.name,
        ip_address=camera_data.ip_address,
        location=camera_data.location,
        is_online=False,
        last_check=datetime.now(timezone.utc)
    )

    db.add(db_camera)
    db.commit()
    db.refresh(db_camera)
    return db_camera


def update_camera(db, camera, camera_data):
    camera.name = camera_data.name
    camera.ip_address = camera_data.ip_address
    camera.location = camera_data.location
    db.commit()
    db.refresh(camera)
    return camera


def scan_camera(db, camera):
    camera.is_online = check_camera_status(camera.ip_address)
    camera.last_check = datetime.now(timezone.utc)
    db.commit()
    db.refresh(camera)
    return camera


def import_cameras_from_excel(db, file_bytes: bytes) -> Tuple[int, int]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows or len(rows) < 1:
        raise ValueError("Excel file is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    field_index = {
        "name": None,
        "ip_address": None,
        "location": None,
        "tên thiết bị": None,
        "địa chỉ ip": None,
        "vị trí lắp đặt": None,
    }

    for idx, header in enumerate(headers):
        if header in field_index:
            field_index[header] = idx

    name_idx = field_index.get("name") if field_index.get("name") is not None else field_index.get("tên thiết bị")
    ip_idx = field_index.get("ip_address") if field_index.get("ip_address") is not None else field_index.get("địa chỉ ip")
    location_idx = field_index.get("location") if field_index.get("location") is not None else field_index.get("vị trí lắp đặt")

    if ip_idx is None or name_idx is None:
        raise ValueError("Excel template must include at least 'name' and 'ip_address' columns")

    created_count = 0
    updated_count = 0
    for row in rows[1:]:
        if not row:
            continue
        ip_value = row[ip_idx] if ip_idx < len(row) else None
        if not ip_value:
            continue
        name_value = row[name_idx] if name_idx < len(row) else None
        location_value = row[location_idx] if location_idx is not None and location_idx < len(row) else None

        ip_address = str(ip_value).strip()
        if not ip_address:
            continue

        existing = db.query(Camera).filter(Camera.ip_address == ip_address).first()
        if existing:
            if name_value is not None:
                existing.name = str(name_value).strip()
            if location_value is not None:
                existing.location = str(location_value).strip()
            updated_count += 1
        else:
            new_camera = Camera(
                name=str(name_value).strip() if name_value is not None else ip_address,
                ip_address=ip_address,
                location=str(location_value).strip() if location_value is not None else None,
                is_online=False,
                last_check=datetime.now(timezone.utc),
            )
            db.add(new_camera)
            created_count += 1

    db.commit()
    return created_count, updated_count


def build_camera_template() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CameraTemplate"
    sheet.append(["name", "ip_address", "location"])
    sheet.append(["Camera 1", "192.168.1.10", "Entrance"])
    sheet.append(["Camera 2", "192.168.1.11", "Lobby"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()